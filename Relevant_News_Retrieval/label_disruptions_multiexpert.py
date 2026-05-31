import os
import json
import time
import random
import threading
from typing import Dict, Any, Tuple, List
from concurrent.futures import ThreadPoolExecutor, as_completed

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

from openai import OpenAI
from google import genai


# ============================
# Models
# ============================
OPENAI_MODEL = "gpt-4.1-mini"
GEMINI_MODEL = "gemini-2.5-flash"

# ============================
# Runtime controls
# ============================
MAX_WORKERS = 20
SAVE_EVERY = 100
MAX_RETRIES = 4
FUTURE_TIMEOUT_S = 240  # 4 mins
JITTER_MAX_S = 0.15

# ============================
# Disruption types
# ============================
TYPES = [
    "flood", "drought", "cyclone_huricane", "extreme_heat", "landslide", "earthquake",
    "mine_accident", "labour_strike", "protests", "trade_embargo",
    "country_relations", "tariffs",
]

SYSTEM_RUBRIC = """
You are labeling news articles for supply-chain disruption signals.

Return a JSON object with ONE 0/1 label for EACH disruption type.

Rules:
- Use 1 ONLY if the article describes that disruption type as an actual event
  affecting operations, production, logistics, trade, or access.
- Use 0 for commentary, background, forecasts, financial reporting, policy discussion
  with no operational impact, historical reference, or unrelated content.
- If unclear or ambiguous, default to 0.

Be conservative: false positives are costly downstream.

Output MUST be valid JSON and match the schema exactly.
""".strip()

def make_user_text(url: str, title: str, meta: str) -> str:
    return (
        f"URL: {url}\n"
        f"TITLE: {title}\n"
        f"META: {meta}\n\n"
        f"Disruption types: {TYPES}\n"
        f"Return JSON with keys: {TYPES}\n"
    )

# ---- OpenAI Structured Outputs (correct format) ----
OPENAI_TEXT_FORMAT = {
    "type": "json_schema",
    "name": "disruption_multilabel",
    "schema": {
        "type": "object",
        "properties": {t: {"type": "integer", "enum": [0, 1]} for t in TYPES},
        "required": TYPES,
        "additionalProperties": False,
    },
    "strict": True,
}


# ============================
# Helpers
# ============================
def _coerce_label(v: Any) -> int:
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)) and v in (0, 1):
        return int(v)
    if isinstance(v, str):
        s = v.strip().lower()
        if s in ("1", "true", "yes"):
            return 1
        if s in ("0", "false", "no"):
            return 0
    raise ValueError(f"Invalid label value: {v!r}")

def validate_payload(payload: Dict[str, Any]) -> Dict[str, int]:
    return {t: _coerce_label(payload[t]) for t in TYPES}

def backoff(attempt: int) -> None:
    time.sleep((2 ** attempt) * 0.6 + random.random() * 0.3)

def strip_fences(text: str) -> str:
    t = (text or "").strip()
    if t.startswith("```"):
        parts = t.split("```")
        if len(parts) >= 3:
            return parts[1].strip()
    return t

def extract_json_object(text: str) -> str:
    """Fallback extractor when the model returns extra text."""
    t = strip_fences(text).strip()
    if not t:
        raise ValueError("Empty model output")
    if t.startswith("{") and t.endswith("}"):
        return t
    start = t.find("{")
    end = t.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError(f"No JSON object found in output: {t[:200]}")
    return t[start:end + 1]


# ============================
# Excel helpers
# ============================
def header_map(ws: Worksheet) -> Dict[str, int]:
    hm: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip():
            hm[v.strip()] = c
    return hm

def ensure_column(ws: Worksheet, hm: Dict[str, int], col_name: str) -> Dict[str, int]:
    """Add a column to the sheet if missing, return updated header map."""
    if col_name in hm:
        return hm
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=col_name)
    hm[col_name] = new_col
    return hm


# ============================
# Thread-local clients
# ============================
_thread_local = threading.local()

def get_openai_client(openai_key: str) -> OpenAI:
    if not hasattr(_thread_local, "openai_client"):
        _thread_local.openai_client = OpenAI(api_key=openai_key)
    return _thread_local.openai_client

def get_gemini_client(gemini_key: str) -> genai.Client:
    if not hasattr(_thread_local, "gemini_client"):
        _thread_local.gemini_client = genai.Client(api_key=gemini_key)
    return _thread_local.gemini_client


# ============================
# Worker
# ============================
def label_row(
    row_id: int,
    url: str,
    title: str,
    meta: str,
    need_cg: bool,
    need_gm: bool,
    openai_key: str,
    gemini_key: str,
) -> Tuple[int, Dict[str, int], Dict[str, int]]:

    time.sleep(random.random() * JITTER_MAX_S)

    openai_client = get_openai_client(openai_key)
    gemini_client = get_gemini_client(gemini_key)

    user_text = make_user_text(url, title, meta)
    cg_out: Dict[str, int] = {}
    gm_out: Dict[str, int] = {}

    # OpenAI (structured outputs)
    if need_cg:
        for a in range(MAX_RETRIES):
            try:
                resp = openai_client.responses.create(
                    model=OPENAI_MODEL,
                    input=[
                        {"role": "system", "content": SYSTEM_RUBRIC},
                        {"role": "user", "content": user_text},
                    ],
                    text={"format": OPENAI_TEXT_FORMAT},
                    temperature=0,
                )
                cg_out = validate_payload(json.loads(resp.output_text))
                break
            except Exception:
                if a == MAX_RETRIES - 1:
                    raise
                backoff(a)

    # Gemini (force JSON)
    if need_gm:
        for a in range(MAX_RETRIES):
            try:
                resp = gemini_client.models.generate_content(
                    model=GEMINI_MODEL,
                    contents=(
                        "Return ONLY a valid JSON object. No markdown. No commentary. No extra text.\n\n"
                        + SYSTEM_RUBRIC + "\n\n" + user_text
                    ),
                    # Key bit: request JSON output
                    config={
                        "response_mime_type": "application/json",
                        "temperature": 0,
                    },
                )

                raw = (resp.text or "").strip()
                # With response_mime_type this should already be JSON,
                # but we keep a fallback extractor for robustness.
                gm_out = validate_payload(json.loads(extract_json_object(raw)))
                break
            except Exception:
                if a == MAX_RETRIES - 1:
                    raise
                backoff(a)

    return row_id, cg_out, gm_out


def main():
    # Load .env from script directory (so it works no matter where you run from)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    load_dotenv(os.path.join(base_dir, ".env"))

    input_xlsx = os.path.join(base_dir, "data", "interim", "disruption_master_10k_multiexpert.xlsx")
    output_xlsx = os.path.join(base_dir, "data", "interim", "disruption_master_10k_multiexpert_labelled.xlsx")
    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)

    openai_key = os.getenv("OPENAI_PROJECT_KEY") or os.getenv("OPENAI_ADMIN_KEY")
    gemini_key = os.getenv("GEMINI_API_KEY")

    if not openai_key:
        raise RuntimeError("Missing OPENAI_PROJECT_KEY or OPENAI_ADMIN_KEY in .env")
    if not gemini_key:
        raise RuntimeError("Missing GEMINI_API_KEY in .env")
    if not os.path.exists(input_xlsx):
        raise FileNotFoundError(f"Input file not found: {input_xlsx}")

    wb = load_workbook(input_xlsx)
    ws = wb["data"]
    hm = header_map(ws)

    # Ensure comparison columns exist
    hm = ensure_column(ws, hm, "agree_all")
    hm = ensure_column(ws, hm, "n_disagree")
    hm = ensure_column(ws, hm, "disagree_labels")

    # Validate required columns
    required_cols = ["row_origin", "url_normalized", "title", "meta_description"]
    for t in TYPES:
        required_cols += [f"chatgpt_{t}", f"gemini_{t}"]
    missing = [c for c in required_cols if c not in hm]
    if missing:
        raise RuntimeError(f"Workbook missing required columns: {missing}")

    # Build jobs
    jobs: List[Tuple[int, str, str, str, bool, bool]] = []
    for r in range(2, ws.max_row + 1):
        origin = ws.cell(r, hm["row_origin"]).value or ""
        if origin == "gold_manual":
            continue

        need_cg = any(ws.cell(r, hm[f"chatgpt_{t}"]).value in (None, "") for t in TYPES)
        need_gm = any(ws.cell(r, hm[f"gemini_{t}"]).value in (None, "") for t in TYPES)
        if not (need_cg or need_gm):
            continue

        url = str(ws.cell(r, hm["url_normalized"]).value or "")
        title = str(ws.cell(r, hm["title"]).value or "")
        meta = str(ws.cell(r, hm["meta_description"]).value or "")
        if not (url or title or meta):
            continue

        jobs.append((r, url, title, meta, need_cg, need_gm))

    print(f"Rows to label: {len(jobs)}")
    if not jobs:
        print("Nothing to do.")
        return

    completed = 0
    failed = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = [
            ex.submit(label_row, r, url, title, meta, need_cg, need_gm, openai_key, gemini_key)
            for (r, url, title, meta, need_cg, need_gm) in jobs
        ]

        pbar = tqdm(total=len(futures), desc="Labelling")
        for fut in as_completed(futures):
            try:
                r, cg_out, gm_out = fut.result(timeout=FUTURE_TIMEOUT_S)
            except Exception as e:
                failed += 1
                # still advance progress so it doesn't "hang"
                pbar.update(1)
                continue

            # Write model outputs
            for t, v in cg_out.items():
                ws.cell(r, hm[f"chatgpt_{t}"], v)
            for t, v in gm_out.items():
                ws.cell(r, hm[f"gemini_{t}"], v)

            # Compare (only if both present)
            if cg_out and gm_out:
                disagree = [t for t in TYPES if int(cg_out[t]) != int(gm_out[t])]
                ws.cell(r, hm["agree_all"], 1 if len(disagree) == 0 else 0)
                ws.cell(r, hm["n_disagree"], len(disagree))
                ws.cell(r, hm["disagree_labels"], ",".join(disagree))
            else:
                # if one model missing, mark as NA
                ws.cell(r, hm["agree_all"], None)
                ws.cell(r, hm["n_disagree"], None)
                ws.cell(r, hm["disagree_labels"], None)

            completed += 1
            if completed % SAVE_EVERY == 0:
                wb.save(output_xlsx)

            pbar.update(1)

        pbar.close()

    wb.save(output_xlsx)
    print(f"Done. Saved to: {output_xlsx}")
    print(f"Completed: {completed} | Failed: {failed}")


if __name__ == "__main__":
    main()